"""Excel export — accuracy first, basic formatting only.

PRODUCT_REQUIREMENTS.md §15 required tabs:
  Summary; Filters Used; Dataset Metadata; Data Quality Warnings; Map Data;
  Geography Summary; Deal vs Portfolio; Market Share; YoY Comparison; Peril;
  Occupancy; Distance to Coast; Geocoding; Stories; Construction; Pivot Output;
  Raw Aggregated Data.

Numbers MUST match the API/screen exactly (CLAUDE.md rule 9). We reuse calc
services that the map/detail/pivot routers already use.

Over `EXPORT_MAX_ROWS` → raise HTTPException(413, EXPORT_TOO_LARGE).
"""

from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import Any

from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..api.exposures import (  # reuse the router builder functions for parity with the wire
    exposures_detail as _build_detail,
    exposures_map as _build_map,
    exposures_pivot as _build_pivot,
)
from ..config import get_settings
from ..models.enums import ErrorCode, Measure, MetricKey, WarningCode
from ..models.exposure import DetailRequest, MapRequest, PivotRequest
from ..models.warnings import make_warning
from ..providers.base import ExposureDataProvider


_HEADER_FILL = PatternFill("solid", fgColor="E0E0E0")
_HEADER_FONT = Font(bold=True)
_TIMESTAMP_FORMAT = "%Y-%m-%dT%H:%M:%SZ"


def _write_header_row(ws, headers: list[str]) -> None:
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = Alignment(horizontal="left")
    ws.freeze_panes = "A2"
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = max(14, len(headers[i - 1]) + 2)


def _write_kv_sheet(wb: Workbook, title: str, rows: list[tuple[str, Any]]) -> None:
    ws = wb.create_sheet(title)
    _write_header_row(ws, ["Field", "Value"])
    for r, (k, v) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=_to_cell(v))


def _to_cell(v: Any) -> Any:
    if v is None:
        return ""
    if isinstance(v, (dict, list, tuple, set)):
        # JSON-ish dump; not pretty but lossless for the cell.
        import json

        return json.dumps(v, default=str, separators=(",", ":"))
    if isinstance(v, datetime):
        return v.strftime(_TIMESTAMP_FORMAT)
    return v


def _check_row_budget(total_rows: int) -> None:
    settings = get_settings()
    if total_rows > settings.export_max_rows:
        raise HTTPException(
            status_code=413,
            detail={
                "code": ErrorCode.EXPORT_TOO_LARGE.value,
                "message": (
                    f"Export would contain {total_rows} rows, exceeding the "
                    f"{settings.export_max_rows}-row limit. Refine filters or aggregation."
                ),
                "details": {"rowCount": total_rows, "limit": settings.export_max_rows},
            },
        )


def build_export_xlsx(payload: dict, provider: ExposureDataProvider) -> bytes:
    """Build the workbook described in PRODUCT_REQUIREMENTS.md §15 and return bytes."""
    timestamp = datetime.now(timezone.utc)

    # ── Compute the analytical responses fresh so numbers match the API exactly ──
    map_req = MapRequest.model_validate(payload, by_alias=True, strict=False)
    map_resp = _build_map(map_req, provider)

    selected_geo = payload.get("selectedGeographyId") or payload.get("selected_geography_id")
    detail_resp = None
    if selected_geo:
        detail_resp = _build_detail(
            DetailRequest.model_validate(
                {**payload, "geographyId": selected_geo}, by_alias=True, strict=False
            ),
            provider,
        )

    pivot_dict = payload.get("pivot") or {}
    pivot_resp = None
    if pivot_dict.get("measures"):
        pivot_req = PivotRequest.model_validate(
            {
                "datasetId": payload.get("datasetId"),
                "datasetGroupId": payload.get("datasetGroupId"),
                "comparisonDatasetId": payload.get("comparisonDatasetId"),
                "filters": payload.get("filters") or {},
                "combinationMethod": payload.get("combinationMethod"),
                **pivot_dict,
            },
            by_alias=True,
            strict=False,
        )
        pivot_resp = _build_pivot(pivot_req, provider)

    # Row-budget guard: rough estimate = map features + breakdown rows + pivot cells + raw facts.
    raw_facts: list = []
    ds_id = payload.get("datasetId")
    if ds_id:
        raw_facts = provider.get_facts_for_dataset(ds_id)
    rough_total = (
        len(map_resp.features)
        + (sum(len(b) for b in detail_resp.breakdowns.model_dump().values()) if detail_resp else 0)
        + (len(pivot_resp.cells) if pivot_resp else 0)
        + len(raw_facts)
    )
    _check_row_budget(rough_total)

    wb = Workbook()
    wb.remove(wb.active)  # drop the default sheet

    # ── Summary ──
    _write_kv_sheet(
        wb,
        "Summary",
        [
            ("Generated at (UTC)", timestamp),
            ("Service", "Exposure Eclipse"),
            ("Dataset ID", payload.get("datasetId")),
            ("Dataset Group ID", payload.get("datasetGroupId")),
            ("Comparison Dataset ID", payload.get("comparisonDatasetId")),
            ("Combination Method", payload.get("combinationMethod")),
            ("Currency", map_resp.currency),
            ("Currency Assumption", payload.get("currencyAssumption")),
            ("Aggregation Level", map_resp.aggregation_level),
            ("Selected Geography", selected_geo),
            ("Metric", map_resp.metric),
            ("Feature Count", len(map_resp.features)),
        ],
    )

    # ── Filters Used ──
    filters = payload.get("filters") or {}
    _write_kv_sheet(wb, "Filters Used", [(k, filters.get(k)) for k in sorted(filters)])

    # ── Dataset Metadata ──
    metadata_rows: list[tuple[str, Any]] = []
    if ds_id:
        prog = provider.get_programme_by_dataset_id(ds_id)
        if prog:
            metadata_rows = list(prog.model_dump(by_alias=True).items())
    _write_kv_sheet(wb, "Dataset Metadata", metadata_rows or [("(none)", "")])

    # ── Data Quality Warnings ──
    ws_warn = wb.create_sheet("Data Quality Warnings")
    _write_header_row(ws_warn, ["Code", "Severity", "Message", "Context"])
    all_warnings = list(map_resp.warnings)
    for f in map_resp.features:
        all_warnings.extend(f.warnings)
    if detail_resp:
        all_warnings.extend(detail_resp.warnings)
    if pivot_resp:
        all_warnings.extend(pivot_resp.warnings)
    for r, w in enumerate(all_warnings, start=2):
        ws_warn.cell(row=r, column=1, value=w.code)
        ws_warn.cell(row=r, column=2, value=w.severity)
        ws_warn.cell(row=r, column=3, value=w.message)
        ws_warn.cell(row=r, column=4, value=_to_cell(w.context))

    # ── Map Data ──
    ws_map = wb.create_sheet("Map Data")
    map_headers = [
        "Geography ID",
        "Geography Name",
        f"Metric ({map_resp.metric})",
        "TIV",
        "Location Count",
        "Deal Share Of Portfolio In Geography",
        "Geography Share Of Total Portfolio",
        "Selected Deal Geography Concentration",
        "Client Market Share",
        "YoY Change",
        "YoY Status",
        "Has Geometry",
    ]
    _write_header_row(ws_map, map_headers)
    for r, f in enumerate(map_resp.features, start=2):
        ws_map.cell(row=r, column=1, value=f.geography_id)
        ws_map.cell(row=r, column=2, value=f.geography_name)
        ws_map.cell(row=r, column=3, value=f.metric_value)
        ws_map.cell(row=r, column=4, value=f.tiv)
        ws_map.cell(row=r, column=5, value=f.location_count)
        ws_map.cell(row=r, column=6, value=f.deal_share_of_portfolio_in_geography)
        ws_map.cell(row=r, column=7, value=f.geography_share_of_total_portfolio)
        ws_map.cell(row=r, column=8, value=f.selected_deal_geography_concentration)
        ws_map.cell(row=r, column=9, value=f.client_market_share)
        ws_map.cell(row=r, column=10, value=f.yoy_change)
        ws_map.cell(row=r, column=11, value=f.yoy_status)
        ws_map.cell(row=r, column=12, value=f.has_geometry)

    # ── Geography Summary (same as map data, for the alias mentioned in the spec) ──
    wb.copy_worksheet(ws_map).title = "Geography Summary"

    # ── Deal vs Portfolio ──
    dvp = detail_resp.deal_vs_portfolio.model_dump() if detail_resp else {}
    _write_kv_sheet(wb, "Deal vs Portfolio", list(dvp.items()) or [("(none)", "")])

    # ── Market Share ──
    ms = detail_resp.market_share.model_dump() if detail_resp else {}
    _write_kv_sheet(wb, "Market Share", list(ms.items()) or [("(none)", "")])

    # ── YoY Comparison ──
    yoy = detail_resp.yoy.model_dump() if detail_resp else {}
    _write_kv_sheet(wb, "YoY Comparison", list(yoy.items()) or [("(none)", "")])

    # ── Breakdowns ──
    breakdown_tabs = {
        "Peril": "peril",
        "Occupancy": "occupancy",
        "Distance to Coast": "distance_to_coast",
        "Geocoding": "geocoding",
        "Stories": "stories",
        "Construction": "construction",
    }
    for sheet, attr in breakdown_tabs.items():
        ws = wb.create_sheet(sheet)
        _write_header_row(ws, ["Key", "TIV", "Location Count"])
        if detail_resp:
            for r, row in enumerate(getattr(detail_resp.breakdowns, attr), start=2):
                ws.cell(row=r, column=1, value=row.key)
                ws.cell(row=r, column=2, value=row.tiv)
                ws.cell(row=r, column=3, value=row.location_count)

    # ── Pivot Output ──
    ws_piv = wb.create_sheet("Pivot Output")
    if pivot_resp:
        headers = [*pivot_resp.rows, *[f"col:{c}" for c in pivot_resp.columns], *pivot_resp.measures]
        _write_header_row(ws_piv, headers or ["(empty)"])
        for r, cell in enumerate(pivot_resp.cells, start=2):
            for ci, part in enumerate(cell.row_key, start=1):
                ws_piv.cell(row=r, column=ci, value=part)
            offset = len(cell.row_key)
            for ci, part in enumerate(cell.col_key, start=1):
                ws_piv.cell(row=r, column=offset + ci, value=part)
            offset += len(cell.col_key)
            for mi, m in enumerate(pivot_resp.measures, start=1):
                ws_piv.cell(row=r, column=offset + mi, value=cell.values.get(m))
    else:
        _write_header_row(ws_piv, ["(no pivot requested)"])

    # ── Raw Aggregated Data ──
    ws_raw = wb.create_sheet("Raw Aggregated Data")
    raw_headers = [
        "Dataset ID",
        "Aggregation",
        "Geography ID",
        "Geography Name",
        "Peril",
        "Occupancy Segment",
        "Construction",
        "Year Built",
        "Distance to Coast",
        "Geocoding Quality",
        "Number of Stories",
        "TIV",
        "Building",
        "Contents",
        "BI",
        "EXPLIM Gross",
        "EXPLIM Net",
        "Location Count",
        "Currency",
    ]
    _write_header_row(ws_raw, raw_headers)
    for r, fact in enumerate(raw_facts, start=2):
        ws_raw.cell(row=r, column=1, value=fact.dataset_id)
        ws_raw.cell(row=r, column=2, value=fact.aggregation)
        ws_raw.cell(row=r, column=3, value=fact.geography_id)
        ws_raw.cell(row=r, column=4, value=fact.state_name or fact.county_name or fact.country_name)
        ws_raw.cell(row=r, column=5, value=fact.peril)
        ws_raw.cell(row=r, column=6, value=fact.occupancy_segment)
        ws_raw.cell(row=r, column=7, value=fact.construction)
        ws_raw.cell(row=r, column=8, value=fact.year_built)
        ws_raw.cell(row=r, column=9, value=fact.distance_to_coast)
        ws_raw.cell(row=r, column=10, value=fact.geocoding_quality)
        ws_raw.cell(row=r, column=11, value=fact.number_of_stories)
        ws_raw.cell(row=r, column=12, value=fact.tiv)
        ws_raw.cell(row=r, column=13, value=fact.building)
        ws_raw.cell(row=r, column=14, value=fact.contents)
        ws_raw.cell(row=r, column=15, value=fact.bi)
        ws_raw.cell(row=r, column=16, value=fact.explim_gross)
        ws_raw.cell(row=r, column=17, value=fact.explim_net)
        ws_raw.cell(row=r, column=18, value=fact.location_count)
        ws_raw.cell(row=r, column=19, value=fact.currency)

    # Surface a top-level warning if export was close to the row limit (config-driven)
    settings = get_settings()
    if rough_total > 0.9 * settings.export_max_rows:
        all_warnings.append(make_warning(WarningCode.WARN_EXPORT_TOO_LARGE))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_hurricane_impact_xlsx(impact: dict) -> bytes:
    """Build a 2-sheet workbook (Summary + Impacted Counties) for a single
    hurricane-impact response. Mirrors the shape returned by
    ``/api/hurricanes/{id}/impact`` exactly — numbers come straight from the
    same compute path used on screen.
    """
    timestamp = datetime.now(timezone.utc)
    wb = Workbook()
    wb.remove(wb.active)

    summary = impact.get("summary") or {}
    counties = impact.get("counties") or []
    _check_row_budget(len(counties))

    _write_kv_sheet(
        wb,
        "Summary",
        [
            ("Generated at (UTC)", timestamp),
            ("Service", "Exposure Eclipse — Hurricane Impact"),
            ("Storm ID", impact.get("stormId")),
            ("Storm Name", impact.get("stormName")),
            ("Year", impact.get("year")),
            ("Currency", impact.get("currency")),
            ("Rmax Multiplier", impact.get("multiplier")),
            ("Counties Impacted", summary.get("countiesImpacted")),
            ("Counties With Portfolio Data", summary.get("countiesWithData")),
            ("Total TIV (in selection)", summary.get("totalTiv")),
            ("Total Location Count", summary.get("totalLocationCount")),
        ],
    )

    ws = wb.create_sheet("Impacted Counties")
    headers = [
        "Geography ID",
        "GEOID (FIPS)",
        "County",
        "State",
        "Max Wind (kt)",
        "Saffir-Simpson",
        "Closest Eye Distance (nm)",
        "Rmax at Closest (nm)",
        "Rmax Source",
        "TIV",
        "Location Count",
        "Has Portfolio Data",
        "Centroid Lat",
        "Centroid Lon",
    ]
    _write_header_row(ws, headers)
    for r, c in enumerate(counties, start=2):
        ws.cell(row=r, column=1, value=c.get("geographyId"))
        ws.cell(row=r, column=2, value=c.get("geoid"))
        ws.cell(row=r, column=3, value=c.get("name"))
        ws.cell(row=r, column=4, value=c.get("state"))
        ws.cell(row=r, column=5, value=c.get("maxWindKt"))
        ws.cell(row=r, column=6, value=c.get("maxCategory"))
        ws.cell(row=r, column=7, value=c.get("closestDistanceNm"))
        ws.cell(row=r, column=8, value=c.get("rmaxAtClosestNm"))
        ws.cell(row=r, column=9, value=c.get("rmaxSource"))
        ws.cell(row=r, column=10, value=c.get("tiv"))
        ws.cell(row=r, column=11, value=c.get("locationCount"))
        ws.cell(row=r, column=12, value=c.get("hasData"))
        ws.cell(row=r, column=13, value=c.get("centroidLat"))
        ws.cell(row=r, column=14, value=c.get("centroidLon"))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# Re-export common types the router may want for convenience.
__all__ = ["build_export_xlsx", "build_hurricane_impact_xlsx", "MetricKey", "Measure"]
