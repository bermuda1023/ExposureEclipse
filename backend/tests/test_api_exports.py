"""Phase 8 — Excel export tests."""

from __future__ import annotations

import io

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.config import get_settings
from app.main import app
from app.providers import get_provider


def setup_function(_):
    get_provider.cache_clear()
    get_settings.cache_clear()


def _payload(**overrides):
    base = {
        "datasetId": "ds-farmers-bda-2027",
        "aggregationLevel": "STATE",
        "metric": "TIV",
        "selectedGeographyId": "US-FL",
        "currency": "USD",
        "pivot": {
            "rows": ["STATE"],
            "columns": ["PERIL"],
            "measures": ["TIV", "LOCATION_COUNT"],
        },
        "filters": {},
    }
    base.update(overrides)
    return base


def test_export_returns_xlsx_with_required_tabs():
    client = TestClient(app)
    resp = client.post("/api/exports/excel", json=_payload())
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    wb = load_workbook(io.BytesIO(resp.content))
    required = {
        "Summary",
        "Filters Used",
        "Dataset Metadata",
        "Data Quality Warnings",
        "Map Data",
        "Geography Summary",
        "Deal vs Portfolio",
        "Market Share",
        "YoY Comparison",
        "Peril",
        "Occupancy",
        "Distance to Coast",
        "Geocoding",
        "Stories",
        "Construction",
        "Pivot Output",
        "Raw Aggregated Data",
    }
    missing = required - set(wb.sheetnames)
    assert not missing, f"Missing tabs: {missing}"


def test_export_map_data_matches_api():
    client = TestClient(app)
    map_resp = client.post(
        "/api/exposures/map",
        json={
            "datasetId": "ds-farmers-bda-2027",
            "aggregationLevel": "STATE",
            "metric": "TIV",
            "filters": {},
        },
    )
    assert map_resp.status_code == 200
    map_features = {f["geographyId"]: f["tiv"] for f in map_resp.json()["features"]}

    export_resp = client.post("/api/exports/excel", json=_payload())
    wb = load_workbook(io.BytesIO(export_resp.content))
    ws = wb["Map Data"]
    sheet_tivs: dict[str, float | None] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        gid = row[0]
        tiv = row[3]
        if gid:
            sheet_tivs[gid] = tiv
    # Every feature in the API response is in the sheet with identical TIV.
    for gid, tiv in map_features.items():
        assert sheet_tivs.get(gid) == tiv, f"Mismatch for {gid}: api={tiv} sheet={sheet_tivs.get(gid)}"


def test_export_too_large_returns_413(monkeypatch):
    # Force a tiny EXPORT_MAX_ROWS so the row-budget guard fires.
    monkeypatch.setenv("EXPORT_MAX_ROWS", "1")
    get_settings.cache_clear()
    client = TestClient(app)
    resp = client.post("/api/exports/excel", json=_payload())
    assert resp.status_code == 413
    body = resp.json()
    assert body["error"]["code"] == "EXPORT_TOO_LARGE"
